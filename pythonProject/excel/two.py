import openpyxl
"""遍历 sheet 1 的所有 数据"""
# 获取 excel对象
ex=openpyxl.load_workbook(r"C:\Users\10756\Desktop\人教版初中英语.xlsx")
sheet1=ex.get_sheet_by_name('七年级上册-Starter Unit 1')
print(sheet1.max_row)
print(sheet1.max_column)
# for x in range(1,sheet1.max_row+1):
#     for y in range(1,sheet1.max_column+1):
#         print(sheet1.cell(row=x,column=y).value,end='\t')
#     print()

# print(openpyxl.utils.get_column_letter(5))
cut=sheet1['A1':'b4']
print(cut[1][1].value)
for row in cut:
    for column in row:
        print(column.value)


