import openpyxl,pprint
# aa={state:{
#     cuntey:{
#         人口:0
#         track:0
#     }
# }}

excelob=openpyxl.load_workbook('censuspopdata.xlsx')
print(excelob.get_sheet_names())
sheet=excelob.get_sheet_by_name('Population by Census Tract')
print(sheet.max_row)
print(sheet.max_column)

censdata={}

for x in range(2,sheet.max_row+1):# 因为 是从第二行才开始是数据
    state=sheet['B'+str(x)].value
    county=sheet['C'+str(x)].value
    pop=sheet['d'+str(x)].value

    # 判断 当前字典里面是否有 这个数据, 如果没有 那么就进行初始化
    censdata.setdefault(state,{})
    censdata[state].setdefault(county,{'pop':0,'track':0})
    # 给当前字典添加数据

    censdata[state][county]['pop']+=int(pop)
    censdata[state][county]['track']+=1

# with open('alldata.py','w') as f:
#     f.write('data='+pprint.pformat(censdata))
total=0

for sta in censdata.keys():
    print(sta+'州有: ' ,end='\t')
    c=0
    for count in censdata[sta]:
        c+=censdata[sta][count]['pop']
    print(str(c)+'人')
    total+=c
print('美国总人数是: '+str(total))